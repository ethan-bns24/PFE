#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genere un classeur Excel structure comme une reference (ex. TPS.xlsx) a
partir de PDF reels (PV d'essai, document commercial).

La reference ne fournit que la structure (noms de feuilles, colonnes,
couleur d'onglet) : aucune valeur n'en est recopiee. Seules les tables
Tb_TPS_Modele (plages dimensionnelles) et Ls_* (listes de valeurs) sont
remplies, a partir de ce qui est effectivement detecte dans les PDF. Les
tables Tc_* (compatibilites) et Tb_Formule (calculs) restent avec leurs
en-tetes seuls : cette logique d'ingenierie n'existe dans aucun PDF.
"""

import re
import unicodedata

import openpyxl
import pdfplumber
from openpyxl.styles.colors import Color

from gamme_excel import feuilles_listes_simples

MOTS_CLES_TABLE_DIMENSIONS = ("largeur", "hauteur", "mini", "maxi", "vantaux", "resistance")
THEME_VERT = 9
TINT_VERT = 0.3999755851924192


def _sans_accents(texte: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", str(texte)) if not unicodedata.combining(c))


def _normaliser_pour_recherche(texte) -> str:
    return re.sub(r"[\s()]+", "", _sans_accents(str(texte))).upper()


def _texte_cellule(valeur) -> str:
    return _sans_accents(str(valeur or "")).lower()


def _parser_nombre(texte):
    texte = re.sub(r"^(mini|maxi)\s*", "", str(texte or "").strip(), flags=re.IGNORECASE)
    texte = texte.replace(" ", " ").replace("\xa0", " ")
    trouve = re.search(r"-?\d[\d ]*(?:[.,]\d+)?", texte)
    if not trouve:
        return None
    nombre = trouve.group(0).replace(" ", "").replace(",", ".")
    return float(nombre) if "." in nombre else int(nombre)


def detecter_tables_dimensions(fichiers_pdf: dict) -> list:
    """fichiers_pdf : {nom_affiche: fichier_ou_chemin}. Retourne les tables
    brutes (liste de lignes) dont l'entete ressemble a un tableau de
    plages dimensionnelles, avec leur provenance."""
    resultats = []
    for nom, fichier in fichiers_pdf.items():
        with pdfplumber.open(fichier) as pdf:
            for i, page in enumerate(pdf.pages):
                for table in page.extract_tables():
                    entete = " ".join(_texte_cellule(c) for row in table[:2] for c in row)
                    nb_motscles = sum(1 for m in MOTS_CLES_TABLE_DIMENSIONS if m in entete)
                    if nb_motscles >= 2:
                        resultats.append({"fichier": nom, "page": i, "table": table})
    return resultats


def _indices_colonnes(table: list) -> dict:
    """Associe un role de colonne a un indice, a partir des 2 premieres
    lignes (en-tetes eventuellement groupes sur 2 lignes). Un groupe
    "Largeur (mm)" / "Hauteur (mm)" issu d'une cellule fusionnee n'occupe
    textuellement que sa colonne de depart : la colonne suivante (sans
    texte propre) est prise comme le Maxi correspondant."""
    entetes = table[:2]
    nb_colonnes = max(len(row) for row in table)
    roles = {}
    for col in range(nb_colonnes):
        blob = " ".join(_texte_cellule(row[col]) for row in entetes if col < len(row))
        if ("type" in blob or "modele" in blob) and "type" not in roles:
            roles["type"] = col
        if ("vantaux" in blob or "vantail" in blob) and "vantaux" not in roles:
            roles["vantaux"] = col
        if ("resistance" in blob or "feu" in blob) and "resistance" not in roles:
            roles["resistance"] = col
        if "largeur" in blob and "largeur_mini" not in roles:
            roles["largeur_mini"], roles["largeur_maxi"] = col, col + 1
        if "hauteur" in blob and "hauteur_mini" not in roles:
            roles["hauteur_mini"], roles["hauteur_maxi"] = col, col + 1
    return roles


def interpreter_table_dimensions(table: list) -> list:
    roles = _indices_colonnes(table)
    if "largeur_mini" not in roles or "hauteur_mini" not in roles:
        return []

    def valeur(row, cle):
        idx = roles.get(cle)
        if idx is None or idx >= len(row) or row[idx] is None:
            return None
        return row[idx]

    lignes = []
    derniere_ligne_modele = None
    for row in table[2:]:
        type_brut = valeur(row, "type")
        vantaux_brut = valeur(row, "vantaux")
        la_min = _parser_nombre(valeur(row, "largeur_mini"))
        la_max = _parser_nombre(valeur(row, "largeur_maxi"))
        ht_min = _parser_nombre(valeur(row, "hauteur_mini"))
        ht_max = _parser_nombre(valeur(row, "hauteur_maxi"))
        resistance = valeur(row, "resistance")

        if not any([type_brut, vantaux_brut, la_min, la_max, ht_min, ht_max]):
            continue

        if type_brut and str(type_brut).strip():
            code = re.sub(r"\s+", "", str(type_brut)).upper()
            ligne = {
                "Ce_Modele": code,
                "Li_Modele": code,
                "Qt_Vantaux": _parser_nombre(vantaux_brut) if vantaux_brut else None,
                "Co_LPL_Min": la_min,
                "Co_LPL_Max": la_max,
                "Co_HPL_Min": ht_min,
                "Co_HPL_Max": ht_max,
                "Ce_ResistanceFeu_detectee": str(resistance).replace("\n", " ").strip() if resistance else None,
                "a_completer": not all(v is not None for v in (la_min, la_max, ht_min, ht_max)),
            }
            lignes.append(ligne)
            derniere_ligne_modele = ligne
        else:
            note = str(vantaux_brut or type_brut or "").replace("\n", " ").strip()
            lignes.append({
                "Ce_Modele": "(a completer)",
                "Li_Modele": note or "variante non identifiee",
                "Qt_Vantaux": derniere_ligne_modele["Qt_Vantaux"] if derniere_ligne_modele else None,
                "Co_LPL_Min": la_min,
                "Co_LPL_Max": la_max,
                "Co_HPL_Min": ht_min,
                "Co_HPL_Max": ht_max,
                "Ce_ResistanceFeu_detectee": None,
                "a_completer": True,
            })
    return lignes


DUREES_FEU_STANDARD = ("15", "20", "30", "45", "60", "90", "120", "180", "240")


def extraire_classements_feu(textes: dict) -> list:
    """Recherche des codes de classement feu (EI60, E120...). Restreint aux
    durees normalisees (15 a 240 min) pour eviter les faux positifs sur des
    numeros de reference ou codes postaux contenant une lettre 'E' isolee."""
    texte_complet = " ".join(t for t in textes.values() if t)
    motif = r"\bEI?\s*\(?\d?\)?\s*(?:" + "|".join(DUREES_FEU_STANDARD) + r")\b"
    codes = set()
    for brut in re.findall(motif, texte_complet, flags=re.IGNORECASE):
        prefixe = "EI" if brut[:2].upper() == "EI" else "E"
        duree = re.search(r"(" + "|".join(DUREES_FEU_STANDARD) + r")$", brut).group(1)
        codes.add(f"{prefixe}{duree}")
    return sorted(codes)


def detecter_mentions_listes(reference_gamme: dict, textes: dict) -> dict:
    texte_complet = _normaliser_pour_recherche(" ".join(t for t in textes.values() if t))
    resultats = {}
    for nom, feuille in feuilles_listes_simples(reference_gamme):
        code_col, label_col = feuille["colonnes"]
        trouves = [
            ligne for ligne in feuille["lignes"]
            if (ligne.get(code_col) and _normaliser_pour_recherche(ligne[code_col]) in texte_complet)
            or (ligne.get(label_col) and _normaliser_pour_recherche(ligne[label_col]) in texte_complet)
        ]
        if trouves:
            resultats[nom] = trouves
    return resultats


def generer_classeur(reference_gamme: dict, lignes_modele: list, mentions_listes: dict):
    classeur = openpyxl.Workbook()
    classeur.remove(classeur.active)

    for nom, feuille in reference_gamme["feuilles"].items():
        ws = classeur.create_sheet(title=feuille["feuille"][:31])
        ws.sheet_properties.tabColor = Color(theme=THEME_VERT, tint=TINT_VERT)
        ws.append([feuille["titre"]])
        ws.append(feuille["colonnes"])
        for cell in ws[2]:
            cell.font = cell.font.copy(bold=True)

        colonnes = feuille["colonnes"]
        if nom == "Tb_TPS_Modele":
            for ligne in lignes_modele:
                ws.append([ligne.get(c) for c in colonnes])
        elif nom in mentions_listes:
            for ligne in mentions_listes[nom]:
                ws.append([ligne.get(c) for c in colonnes])

    return classeur
